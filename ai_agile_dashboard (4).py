import streamlit as st
import pandas as pd
import numpy as np
from sklearn.linear_model import LogisticRegression, LinearRegression
from sklearn.ensemble import RandomForestClassifier
from sklearn.model_selection import train_test_split
from sklearn.metrics import accuracy_score, classification_report, mean_squared_error
from sklearn.preprocessing import LabelEncoder

st.set_page_config(page_title="AI Agile Dashboard", layout="wide")

st.markdown("""
<style>
    .agent-card {
        background: linear-gradient(135deg, #1e1e2e 0%, #2a2a3e 100%);
        border: 1px solid #444466;
        border-radius: 12px;
        padding: 1.2rem 1.5rem;
        margin-bottom: 1rem;
        color: #e0e0f0;
    }
    .agent-card.critical {
        border-left: 5px solid #ff4d6d;
        background: linear-gradient(135deg, #2e1e22 0%, #3e2a2e 100%);
    }
    .agent-card.warning {
        border-left: 5px solid #ffd166;
        background: linear-gradient(135deg, #2e2a1e 0%, #3e362a 100%);
    }
    .agent-card.success {
        border-left: 5px solid #06d6a0;
        background: linear-gradient(135deg, #1e2e2a 0%, #2a3e36 100%);
    }
    .agent-card.info {
        border-left: 5px solid #4cc9f0;
        background: linear-gradient(135deg, #1e252e 0%, #2a333e 100%);
    }
    .agent-title {
        font-size: 1rem;
        font-weight: 700;
        margin-bottom: 0.3rem;
        letter-spacing: 0.03em;
    }
    .agent-detail {
        font-size: 0.85rem;
        opacity: 0.85;
        line-height: 1.5;
    }
    .chain-step {
        display: flex;
        align-items: flex-start;
        gap: 1rem;
        margin-bottom: 0.8rem;
    }
    .step-num {
        background: #4cc9f0;
        color: #000;
        border-radius: 50%;
        width: 28px;
        height: 28px;
        display: flex;
        align-items: center;
        justify-content: center;
        font-weight: 800;
        font-size: 0.8rem;
        flex-shrink: 0;
    }
    .health-bar-container {
        background: #333;
        border-radius: 8px;
        height: 14px;
        width: 100%;
        margin-top: 4px;
    }
    .report-section {
        background: #1a1a2e;
        border-radius: 10px;
        padding: 1rem 1.5rem;
        margin-bottom: 1rem;
        border: 1px solid #333355;
        color: #dde;
        font-size: 0.9rem;
        line-height: 1.7;
    }
</style>
""", unsafe_allow_html=True)

st.title("🚀 AI Agile Project Management Dashboard")

uploaded_file = st.file_uploader("📁 Upload the Combined CSV for All Objectives", type="csv")

# ── shared state ────────────────────────────────────────────────────────────
models = {}   # will hold trained models keyed by objective
encoders = {} # label encoders

if uploaded_file:
    df = pd.read_csv(uploaded_file)
    df = df.fillna(0)

    # Handle both string ('Yes'/'No') and float/probability labels
    def binarize_col(series, threshold=0.5):
        if series.dtype == object:
            return series.map({'No': 0, 'Yes': 1}).fillna(0).astype(int)
        return (series > threshold).astype(int)

    thresholds = {'Success_Label': 0.5, 'Expected_Overload': 0.5, 'Risk_Flag': 0.3}
    for col, thresh in thresholds.items():
        if col in df.columns:
            df[col] = binarize_col(df[col], threshold=thresh)

    st.success("✅ File uploaded successfully!")

    # Dataset summary
    c1, c2, c3, c4 = st.columns(4)
    c1.metric("📋 Total Records", f"{len(df):,}")
    c2.metric("📊 Features", len(df.columns))
    if 'Success_Label' in df.columns:
        c3.metric("🔴 Sprints at Risk", int((df['Success_Label']==0).sum()))
    if 'Risk_Flag' in df.columns:
        c4.metric("⚠️ Burnout Flags", int(df['Risk_Flag'].sum()))

    with st.expander("👀 Preview Data"):
        st.write(df.head())

    # ── train all models silently so the agent can use them ─────────────────
    def train_all(df):
        results = {}

        # --- Obj 1: Sprint Completion ---
        try:
            X1 = df[['Planned_Story_Points_Sprint','Completed_Story_Points','Percent_Done',
                      'Days_Remaining_Sprint','Historical_Velocity','Blocked_Stories','Scope_Change']]
            y1 = df['Success_Label']
            if len(y1.unique()) > 1:
                m = LogisticRegression(max_iter=1000, class_weight='balanced')
                m.fit(X1, y1)
                results['sprint'] = {'model': m, 'features': X1.columns.tolist()}
        except: pass

        # --- Obj 2: Workload ---
        try:
            X2 = df[['Planned_Story_Points_Resource','Current_Assigned_SP','Historical_Avg_SP',
                      'Remaining_Days_Resource','High_Priority_Tasks_Resource','Current_Workload_Percent']]
            y2 = df['Expected_Overload']
            if len(y2.unique()) > 1:
                m = RandomForestClassifier(n_estimators=100, random_state=42)
                m.fit(X2, y2)
                results['workload'] = {'model': m, 'features': X2.columns.tolist()}
        except: pass

        # --- Obj 3: Time to Resolve ---
        try:
            X3 = pd.get_dummies(df[['Issue_Type','Priority']], drop_first=False)
            X3 = pd.concat([X3, df[['Original_Estimate_Hours','Story_Points_Issue']]], axis=1)
            y3 = df['Resolution_Time_Hours']
            m = LinearRegression()
            m.fit(X3, y3)
            results['ttr'] = {'model': m, 'features': X3.columns.tolist(), 'X3': X3}
        except: pass

        # --- Obj 4: Burnout ---
        try:
            X4 = df[['Total_SP_This_Sprint','Historical_Avg_SP_Burnout',
                      'High_Priority_Tasks_Burnout','Consecutive_Overloads']]
            y4 = df['Risk_Flag']
            if len(y4.unique()) > 1:
                m = RandomForestClassifier(n_estimators=100, random_state=42)
                m.fit(X4, y4)
                results['burnout'] = {'model': m, 'features': X4.columns.tolist()}
        except: pass

        # --- Obj 5: Resource Allocation ---
        try:
            le_s = LabelEncoder(); le_l = LabelEncoder()
            df['Summary_enc'] = le_s.fit_transform(df['Summary'].astype(str))
            df['Labels_enc']  = le_l.fit_transform(df['Labels'].astype(str))
            X5 = df[['Summary_enc','Labels_enc','Original_Estimate_Resource','Story_Points_Resource']]
            y5 = df['Assignee_Resource']
            m = RandomForestClassifier(n_estimators=100, random_state=42)
            m.fit(X5, y5)
            results['alloc'] = {'model': m, 'features': X5.columns.tolist(),
                                'le_summary': le_s, 'le_labels': le_l}
        except: pass

        return results, df

    models, df = train_all(df)

    # ── agentic scan: run all models on every row ────────────────────────────
    def run_agent_scan(_df, _models_keys):
        """Run predictions across the whole dataset and collect findings."""
        findings = []
        df = _df.copy()

        # Sprint risk scan — aggregate, not per-row
        if 'sprint' in models:
            m = models['sprint']['model']
            cols = models['sprint']['features']
            try:
                X = df[cols]
                preds  = m.predict(X)
                probas = m.predict_proba(X)[:, 1]
                at_risk_mask = (preds == 0)
                at_risk_count = int(at_risk_mask.sum())
                total = len(preds)
                pct = at_risk_count / total if total > 0 else 0
                if pct > 0.15:
                    avg_prob   = float(probas[at_risk_mask].mean())
                    avg_blocked = float(df.loc[at_risk_mask, 'Blocked_Stories'].mean()) if 'Blocked_Stories' in df.columns else 0
                    avg_days    = float(df.loc[at_risk_mask, 'Days_Remaining_Sprint'].mean()) if 'Days_Remaining_Sprint' in df.columns else 0
                    sev = 'critical' if pct > 0.5 else 'warning'
                    findings.append({
                        'severity': sev,
                        'objective': 'Sprint Completion',
                        'icon': '🔴' if sev == 'critical' else '🟡',
                        'title': f"{at_risk_count} of {total} sprints ({pct:.0%}) at risk of spillover",
                        'detail': (f"Avg completion probability: {avg_prob:.0%} | "
                                   f"Avg blocked stories: {avg_blocked:.1f} | "
                                   f"Avg days remaining: {avg_days:.1f}"),
                        'action': "Consider reducing scope or unblocking stories immediately."
                    })
            except: pass

        # Workload overload scan
        if 'workload' in models:
            m = models['workload']['model']
            cols = models['workload']['features']
            try:
                X = df[cols]
                preds  = m.predict(X)
                probas = m.predict_proba(X)[:, 1]
                overloaded = df[preds == 1].copy()
                overloaded['wl_prob'] = probas[preds == 1]
                count = len(overloaded)
                if count > 0 and count > len(df) * 0.2:
                    findings.append({
                        'severity': 'critical' if count > len(df) * 0.45 else 'warning',
                        'objective': 'Workload Projection',
                        'icon': '🔴' if count > len(df) * 0.45 else '🟡',
                        'title': f"{count} resource(s) projected to be overloaded",
                        'detail': (f"Average overload probability: {overloaded['wl_prob'].mean():.0%} | "
                                   f"Avg current workload: {overloaded.get('Current_Workload_Percent', pd.Series([0])).mean():.0f}%"),
                        'action': "Redistribute story points from overloaded to available team members."
                    })
            except: pass

        # Burnout risk scan
        if 'burnout' in models:
            m = models['burnout']['model']
            cols = models['burnout']['features']
            try:
                X = df[cols]
                preds = m.predict(X)
                at_risk_count = int(preds.sum())
                pct_flagged_b = at_risk_count / len(preds) if len(preds) > 0 else 0
                if pct_flagged_b > 0.25:
                    avg_co = df.loc[preds == 1, 'Consecutive_Overloads'].mean() if 'Consecutive_Overloads' in df.columns else 0
                    pct_flagged = pct_flagged_b
                    findings.append({
                        'severity': 'critical' if pct_flagged > 0.5 else 'warning',
                        'objective': 'Burnout Risk',
                        'icon': '🔴' if pct_flagged > 0.5 else '🟡',
                        'title': f"{at_risk_count} team member(s) flagged for burnout risk",
                        'detail': f"Avg consecutive overloads: {avg_co:.1f} sprints",
                        'action': "Schedule 1:1s, reduce high-priority task load, or grant recovery sprint."
                    })
            except: pass

        # Healthy signal
        sprint_ok   = sum(1 for f in findings if f['objective'] == 'Sprint Completion') == 0
        workload_ok = sum(1 for f in findings if f['objective'] == 'Workload Projection') == 0
        burnout_ok  = sum(1 for f in findings if f['objective'] == 'Burnout Risk') == 0

        if sprint_ok:
            findings.append({'severity':'success','objective':'Sprint Completion','icon':'✅',
                             'title':'All sprints on track','detail':'No spillover risk detected.',
                             'action':''})
        if workload_ok:
            findings.append({'severity':'success','objective':'Workload Projection','icon':'✅',
                             'title':'Workloads within capacity','detail':'No overload signals found.',
                             'action':''})
        if burnout_ok:
            findings.append({'severity':'success','objective':'Burnout Risk','icon':'✅',
                             'title':'No burnout risk detected','detail':'Team load looks sustainable.',
                             'action':''})

        return findings

    # ── chained decisions ────────────────────────────────────────────────────
    def build_chain(findings):
        """Build an agentic chain of decisions based on cross-objective findings."""
        chain = []
        has_sprint_risk   = any(f['objective'] == 'Sprint Completion'  and f['severity'] in ('critical','warning') for f in findings)
        has_overload      = any(f['objective'] == 'Workload Projection' and f['severity'] in ('critical','warning') for f in findings)
        has_burnout       = any(f['objective'] == 'Burnout Risk'        and f['severity'] in ('critical','warning') for f in findings)

        chain.append({
            'step': 1,
            'label': 'Scan Objectives',
            'detail': f"Scanned {len(df)} records across 5 objectives.",
            'status': 'done'
        })

        if has_sprint_risk and has_overload:
            chain.append({'step':2,'label':'Linked: Sprint risk ← Overload detected',
                'detail':'Sprint may be at risk because team members are overloaded. Reallocation needed.',
                'status':'alert'})
            chain.append({'step':3,'label':'Recommend: Rebalance workload',
                'detail':'Move story points from overloaded members to those under capacity before sprint closes.',
                'status':'action'})
        elif has_sprint_risk:
            chain.append({'step':2,'label':'Sprint risk detected — checking workload',
                'detail':'Workload looks OK. Risk may stem from blocked stories or scope change.',
                'status':'alert'})
            chain.append({'step':3,'label':'Recommend: Unblock stories & freeze scope',
                'detail':'No reallocation needed. Focus on removing blockers and preventing scope creep.',
                'status':'action'})

        if has_burnout and has_overload:
            chain.append({'step': len(chain)+1,'label':'Linked: Burnout risk ← Persistent overloads',
                'detail':'Burnout signal correlates with overloaded workload across multiple sprints.',
                'status':'alert'})
            chain.append({'step': len(chain)+1,'label':'Recommend: Recovery sprint or capacity reduction',
                'detail':'Reduce assigned story points by 20–30% for flagged members next sprint.',
                'status':'action'})

        if not has_sprint_risk and not has_overload and not has_burnout:
            chain.append({'step':2,'label':'All clear — no chained risks',
                'detail':'No cross-objective dependencies triggered. Project health looks good.',
                'status':'done'})

        return chain

    # ── health score ─────────────────────────────────────────────────────────
    def compute_health(findings):
        score = 100
        for f in findings:
            if f['severity'] == 'critical': score -= 25
            elif f['severity'] == 'warning': score -= 10
        return max(0, min(100, score))

    # ── generate written report ───────────────────────────────────────────────
    def generate_report(findings, chain, score):
        total   = len(df)
        criticals = [f for f in findings if f['severity'] == 'critical']
        warnings  = [f for f in findings if f['severity'] == 'warning']
        successes = [f for f in findings if f['severity'] == 'success']

        status = "🟢 Healthy" if score >= 75 else ("🟡 Needs Attention" if score >= 50 else "🔴 At Risk")

        report = f"""
## 📋 Project Health Report

**Overall Status:** {status} — Health Score: {score}/100

**Dataset:** {total} records analyzed across 5 AI objectives.

### Summary
The autonomous agent scanned your project data and identified **{len(criticals)} critical issue(s)** and **{len(warnings)} warning(s)**. {len(successes)} objective(s) returned healthy signals.

### Findings
"""
        for f in findings:
            if f['severity'] != 'success':
                report += f"\n- **{f['icon']} [{f['objective']}]** {f['title']}: {f['detail']}"
                if f['action']:
                    report += f"\n  → *{f['action']}*"

        report += "\n\n### Healthy Signals\n"
        for f in successes:
            report += f"\n- **{f['icon']} [{f['objective']}]** {f['title']}"

        report += "\n\n### Chained Recommendations\n"
        for step in chain:
            emoji = "✅" if step['status'] == 'done' else ("⚠️" if step['status'] == 'alert' else "💡")
            report += f"\n**Step {step['step']}** {emoji} {step['label']}  \n{step['detail']}\n"

        report += f"\n\n---\n*Report auto-generated by the Agentic AI layer. {total} rows × 5 objectives scanned.*"
        return report

    # ═══════════════════════════════════════════════════════════════════════
    tabs = st.tabs([
        "🤖 Agentic AI Overview",
        "1️⃣ Sprint Completion Forecast",
        "2️⃣ Workload Projection Forecast",
        "3️⃣ Time to Resolve Estimation",
        "4️⃣ Burnout Risk Alerts",
        "5️⃣ Resource Allocation Suggestions"
    ])

    # ══════════════════════════════════════════════════════════════
    # AGENTIC AI TAB
    # ══════════════════════════════════════════════════════════════
    with tabs[0]:
        st.header("🤖 Agentic AI — Autonomous Project Scanning")
        st.caption("The agent automatically runs all 5 models across your full dataset, chains findings together, and surfaces prioritized actions — no manual input needed.")

        with st.spinner("🧠 Agent scanning dataset across all objectives..."):
            findings = run_agent_scan(df, list(models.keys()))
            chain    = build_chain(findings)
            score    = compute_health(findings)

        # ── Health Score ────────────────────────────────────────
        st.markdown("---")
        col1, col2, col3 = st.columns([1, 2, 1])
        with col1:
            color = "#06d6a0" if score >= 75 else ("#ffd166" if score >= 50 else "#ff4d6d")
            label = "🟢 Healthy" if score >= 75 else ("🟡 Needs Attention" if score >= 50 else "🔴 At Risk")
            st.markdown(f"""
            <div style='text-align:center;'>
                <div style='font-size:3.5rem;font-weight:900;color:{color};'>{score}</div>
                <div style='font-size:1rem;color:#aaa;margin-top:-8px;'>/ 100</div>
                <div style='font-size:1.1rem;margin-top:6px;'>{label}</div>
                <div style='font-size:0.8rem;color:#888;'>Project Health Score</div>
            </div>
            """, unsafe_allow_html=True)
        with col2:
            bar_color = "#06d6a0" if score >= 75 else ("#ffd166" if score >= 50 else "#ff4d6d")
            st.markdown(f"""
            <div style='margin-top:2rem;'>
                <div class='health-bar-container'>
                    <div style='background:{bar_color};width:{score}%;height:14px;border-radius:8px;transition:width 1s ease;'></div>
                </div>
                <div style='display:flex;justify-content:space-between;font-size:0.75rem;color:#888;margin-top:4px;'>
                    <span>0 — Critical</span><span>50 — Attention</span><span>100 — Healthy</span>
                </div>
            </div>
            """, unsafe_allow_html=True)
        with col3:
            criticals_n = sum(1 for f in findings if f['severity']=='critical')
            warnings_n  = sum(1 for f in findings if f['severity']=='warning')
            st.markdown(f"""
            <div style='text-align:center;margin-top:0.5rem;'>
                <div style='font-size:2rem;font-weight:800;color:#ff4d6d;'>{criticals_n}</div>
                <div style='font-size:0.8rem;color:#aaa;'>Critical Issues</div>
                <div style='font-size:2rem;font-weight:800;color:#ffd166;margin-top:8px;'>{warnings_n}</div>
                <div style='font-size:0.8rem;color:#aaa;'>Warnings</div>
            </div>
            """, unsafe_allow_html=True)

        st.markdown("---")

        # ── Findings ────────────────────────────────────────────
        st.subheader("🔍 Autonomous Findings")
        severity_order = {'critical': 0, 'warning': 1, 'info': 2, 'success': 3}
        for f in sorted(findings, key=lambda x: severity_order.get(x['severity'], 99)):
            action_html = f"<div style='margin-top:6px;font-style:italic;opacity:0.75;'>→ {f['action']}</div>" if f['action'] else ""
            st.markdown(f"""
            <div class='agent-card {f["severity"]}'>
                <div class='agent-title'>{f["icon"]} [{f["objective"]}] {f["title"]}</div>
                <div class='agent-detail'>{f["detail"]}{action_html}</div>
            </div>
            """, unsafe_allow_html=True)

        st.markdown("---")

        # ── Decision Chain ──────────────────────────────────────
        st.subheader("⛓️ Chained Decision Reasoning")
        st.caption("The agent links findings across objectives to produce connected, prioritized recommendations.")

        for step in chain:
            icon   = "✅" if step['status'] == 'done' else ("⚠️" if step['status'] == 'alert' else "💡")
            color  = "#4cc9f0" if step['status'] == 'done' else ("#ffd166" if step['status'] == 'alert' else "#06d6a0")
            sev    = "info" if step['status'] == 'done' else ("warning" if step['status'] == 'alert' else "success")
            st.markdown(f"""
            <div class='agent-card {sev}' style='display:flex;gap:1rem;align-items:flex-start;'>
                <div style='background:{color};color:#000;border-radius:50%;width:30px;height:30px;
                            display:flex;align-items:center;justify-content:center;
                            font-weight:800;font-size:0.8rem;flex-shrink:0;margin-top:2px;'>
                    {step["step"]}
                </div>
                <div>
                    <div class='agent-title'>{icon} {step["label"]}</div>
                    <div class='agent-detail'>{step["detail"]}</div>
                </div>
            </div>
            """, unsafe_allow_html=True)

        st.markdown("---")

        # ── Per-Assignee Breakdown ──────────────────────────────
        st.subheader("👤 Per-Assignee Risk Breakdown")
        st.caption("Risk levels computed directly from your dataset for each team member.")

        assignee_col = 'Assignee' if 'Assignee' in df.columns else ('Assignee_Resource' if 'Assignee_Resource' in df.columns else None)
        if assignee_col:
            assignees = df[assignee_col].unique()
            cols_a = st.columns(len(assignees))
            for i, person in enumerate(sorted(assignees)):
                sub = df[df[assignee_col] == person]
                sprint_risk = sub['Success_Label'].eq(0).mean() if 'Success_Label' in df.columns else 0
                overload    = sub['Expected_Overload'].mean()   if 'Expected_Overload' in df.columns else 0
                burnout     = sub['Risk_Flag'].mean()           if 'Risk_Flag' in df.columns else 0
                workload    = sub['Current_Workload_Percent'].mean() if 'Current_Workload_Percent' in df.columns else 0
                consec      = sub['Consecutive_Overloads'].mean()    if 'Consecutive_Overloads' in df.columns else 0

                # Overall person score
                person_score = 100 - (sprint_risk * 35) - (overload * 30) - (burnout * 20) - min((workload - 100) / 2, 15)
                person_score = max(0, min(100, person_score))
                p_color = "#06d6a0" if person_score >= 60 else ("#ffd166" if person_score >= 40 else "#ff4d6d")
                p_label = "🟢 OK" if person_score >= 60 else ("🟡 Watch" if person_score >= 40 else "🔴 At Risk")

                with cols_a[i]:
                    st.markdown(f"""
                    <div class='agent-card {"critical" if person_score < 40 else "warning" if person_score < 60 else "success"}' style='text-align:center;'>
                        <div style='font-size:1.5rem;font-weight:900;color:{p_color};'>{person_score:.0f}</div>
                        <div style='font-size:0.7rem;color:#aaa;margin-top:-4px;'>/ 100</div>
                        <div style='font-size:1rem;font-weight:700;margin:6px 0 2px;'>{person}</div>
                        <div style='font-size:0.75rem;margin-bottom:8px;'>{p_label}</div>
                        <div style='text-align:left;font-size:0.78rem;line-height:1.8;'>
                            🏃 Sprint risk: <b>{sprint_risk:.0%}</b><br>
                            📦 Overload: <b>{overload:.0%}</b><br>
                            🔥 Burnout flag: <b>{burnout:.0%}</b><br>
                            ⚡ Avg workload: <b>{workload:.0f}%</b><br>
                            🔁 Consec. overloads: <b>{consec:.1f}</b>
                        </div>
                    </div>
                    """, unsafe_allow_html=True)

        st.markdown("---")

        # ── Action Priority Table ───────────────────────────────
        st.subheader("🎯 Action Priority Table")
        st.caption("Ranked actions based on severity and impact across all findings.")

        action_rows = []
        for f in sorted(findings, key=lambda x: {'critical': 0, 'warning': 1, 'success': 2}.get(x['severity'], 3)):
            if f.get('action'):
                priority = "🔴 P1 — Immediate" if f['severity'] == 'critical' else "🟡 P2 — This Sprint"
                action_rows.append({
                    'Priority':   priority,
                    'Objective':  f['objective'],
                    'Issue':      f['title'],
                    'Action':     f['action']
                })

        # Add chained actions
        for step in chain:
            if step['status'] == 'action':
                action_rows.append({
                    'Priority':  '💡 P3 — Next Sprint',
                    'Objective': 'Cross-Objective',
                    'Issue':     step['label'],
                    'Action':    step['detail']
                })

        if action_rows:
            action_df = pd.DataFrame(action_rows)
            st.dataframe(
                action_df,
                use_container_width=True,
                hide_index=True,
                column_config={
                    'Priority':  st.column_config.TextColumn('Priority',  width='medium'),
                    'Objective': st.column_config.TextColumn('Objective', width='medium'),
                    'Issue':     st.column_config.TextColumn('Issue',     width='large'),
                    'Action':    st.column_config.TextColumn('Action ▶',  width='large'),
                }
            )

        st.markdown("---")

        # ── Written Report ──────────────────────────────────────
        st.subheader("📄 Auto-Generated Project Health Report")
        report_md = generate_report(findings, chain, score)
        st.markdown(f"<div class='report-section'>{report_md}</div>", unsafe_allow_html=True)

        col_dl, col_xl, _ = st.columns([1, 1, 2])
        with col_dl:
            st.download_button(
                "⬇️ Download Report (.md)",
                data=report_md,
                file_name="project_health_report.md",
                mime="text/markdown"
            )

        st.markdown("---")

        # ── Trend Charts ────────────────────────────────────────
        st.subheader("📈 Trend Charts")
        st.caption("Sprint risk and workload trends over time, grouped by sprint number.")

        sprint_col = 'Sprint_Number' if 'Sprint_Number' in df.columns else None
        if sprint_col:
            df_trend = df.copy()
            df_trend['at_risk'] = (df_trend['Success_Label'] == 0).astype(int)
            df_trend['overloaded'] = (df_trend['Expected_Overload'] > 0.5).astype(int) if df_trend['Expected_Overload'].dtype != int else df_trend['Expected_Overload']

            trend_agg = df_trend.groupby(sprint_col).agg(
                sprint_risk_pct=('at_risk', 'mean'),
                avg_workload=('Current_Workload_Percent', 'mean'),
                avg_blocked=('Blocked_Stories', 'mean'),
                burnout_pct=('Risk_Flag', 'mean'),
            ).reset_index()
            trend_agg['sprint_risk_pct'] = (trend_agg['sprint_risk_pct'] * 100).round(1)
            trend_agg['avg_workload'] = trend_agg['avg_workload'].round(1)
            trend_agg['avg_blocked'] = trend_agg['avg_blocked'].round(2)
            trend_agg['burnout_pct'] = (trend_agg['burnout_pct'] * 100).round(1)
            trend_agg = trend_agg.sort_values(sprint_col)

            tc1, tc2 = st.columns(2)
            with tc1:
                st.markdown("**🏃 Sprint Risk % Over Time**")
                st.line_chart(trend_agg.set_index(sprint_col)[['sprint_risk_pct']], height=220, use_container_width=True)
            with tc2:
                st.markdown("**⚡ Avg Workload % Over Time**")
                st.line_chart(trend_agg.set_index(sprint_col)[['avg_workload']], height=220, use_container_width=True)

            tc3, tc4 = st.columns(2)
            with tc3:
                st.markdown("**🚧 Avg Blocked Stories Over Time**")
                st.line_chart(trend_agg.set_index(sprint_col)[['avg_blocked']], height=220, use_container_width=True)
            with tc4:
                st.markdown("**🔥 Burnout Flag % Over Time**")
                st.line_chart(trend_agg.set_index(sprint_col)[['burnout_pct']], height=220, use_container_width=True)

            st.markdown("**👤 Workload Trend per Assignee**")
            assignee_col_t = 'Assignee' if 'Assignee' in df.columns else None
            if assignee_col_t:
                wl_pivot = df_trend.groupby([sprint_col, assignee_col_t])['Current_Workload_Percent'].mean().unstack(assignee_col_t).round(1)
                wl_pivot = wl_pivot.sort_index()
                st.line_chart(wl_pivot, height=280, use_container_width=True)
        else:
            st.info("ℹ️ Upload a dataset with a 'Sprint_Number' column to enable trend charts.")

        st.markdown("---")

        # ── Excel Export ─────────────────────────────────────────
        st.subheader("📥 Export to Excel")
        st.caption("Download a full Excel workbook with findings, action table, assignee breakdown, and raw data.")

        try:
            import io
            from openpyxl import Workbook
            from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
            from openpyxl.utils import get_column_letter

            def build_excel(df, findings, chain, score, action_rows):
                wb = Workbook()

                # ---- Sheet 1: Summary ----
                ws1 = wb.active
                ws1.title = "Health Summary"
                header_fill  = PatternFill("solid", fgColor="1e1e2e")
                red_fill     = PatternFill("solid", fgColor="ff4d6d")
                yellow_fill  = PatternFill("solid", fgColor="ffd166")
                green_fill   = PatternFill("solid", fgColor="06d6a0")
                white_font   = Font(color="FFFFFF", bold=True, size=12)
                dark_font    = Font(color="1e1e2e", bold=True, size=11)
                thin_border  = Border(
                    left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin'))

                ws1['A1'] = "AI Agile Dashboard — Project Health Report"
                ws1['A1'].font = Font(bold=True, size=16, color="1e1e2e")
                ws1['A2'] = f"Health Score: {score}/100"
                ws1['A2'].font = Font(bold=True, size=13,
                    color="FF0000" if score < 50 else ("FF9900" if score < 75 else "009900"))
                ws1['A3'] = f"Records analyzed: {len(df):,}   |   Objectives: 5"
                ws1['A3'].font = Font(size=11)
                ws1.append([])

                ws1.append(["Objective", "Severity", "Finding", "Action"])
                for cell in ws1[5]:
                    cell.fill = header_fill
                    cell.font = white_font
                    cell.alignment = Alignment(horizontal='center')

                for f in findings:
                    fill = red_fill if f['severity']=='critical' else (yellow_fill if f['severity']=='warning' else green_fill)
                    font = dark_font
                    row = [f['objective'], f['severity'].upper(), f['title'], f.get('action','')]
                    ws1.append(row)
                    for cell in ws1[ws1.max_row]:
                        cell.fill = fill
                        cell.font = font
                        cell.border = thin_border
                        cell.alignment = Alignment(wrap_text=True)

                for col in ['A','B','C','D']:
                    ws1.column_dimensions[col].width = 28
                ws1.row_dimensions[5].height = 20

                # ---- Sheet 2: Action Priority ----
                ws2 = wb.create_sheet("Action Priority")
                ws2.append(["Priority", "Objective", "Issue", "Recommended Action"])
                for cell in ws2[1]:
                    cell.fill = header_fill
                    cell.font = white_font
                for ar in action_rows:
                    ws2.append([ar.get('Priority',''), ar.get('Objective',''),
                                ar.get('Issue',''), ar.get('Action','')])
                    for cell in ws2[ws2.max_row]:
                        cell.border = thin_border
                        cell.alignment = Alignment(wrap_text=True)
                for col in ['A','B','C','D']:
                    ws2.column_dimensions[col].width = 30

                # ---- Sheet 3: Assignee Breakdown ----
                ws3 = wb.create_sheet("Assignee Breakdown")
                assignee_col = 'Assignee' if 'Assignee' in df.columns else None
                if assignee_col:
                    ws3.append(["Assignee","Health Score","Sprint Risk %","Overload %",
                                 "Burnout Flag %","Avg Workload %","Avg Consec. Overloads","Status"])
                    for cell in ws3[1]:
                        cell.fill = header_fill; cell.font = white_font
                    for person in sorted(df[assignee_col].unique()):
                        sub = df[df[assignee_col]==person]
                        sprint_risk = sub['Success_Label'].eq(0).mean() if 'Success_Label' in df.columns else 0
                        overload    = sub['Expected_Overload'].mean()   if 'Expected_Overload' in df.columns else 0
                        burnout     = sub['Risk_Flag'].mean()           if 'Risk_Flag' in df.columns else 0
                        workload    = sub['Current_Workload_Percent'].mean() if 'Current_Workload_Percent' in df.columns else 0
                        consec      = sub['Consecutive_Overloads'].mean()    if 'Consecutive_Overloads' in df.columns else 0
                        p_score = max(0,min(100,100-(sprint_risk*35)-(overload*30)-(burnout*20)-min((workload-100)/2,15)))
                        status = "OK" if p_score >= 60 else ("Watch" if p_score >= 40 else "At Risk")
                        ws3.append([person, round(p_score,1), f"{sprint_risk:.0%}", f"{overload:.0%}",
                                     f"{burnout:.0%}", f"{workload:.0f}%", round(consec,1), status])
                        fill = green_fill if p_score>=60 else (yellow_fill if p_score>=40 else red_fill)
                        for cell in ws3[ws3.max_row]:
                            cell.fill = fill; cell.font = dark_font; cell.border = thin_border
                    for col in ['A','B','C','D','E','F','G','H']:
                        ws3.column_dimensions[col].width = 22

                # ---- Sheet 4: Raw Data ----
                ws4 = wb.create_sheet("Raw Data")
                cols = df.columns.tolist()
                ws4.append(cols)
                for cell in ws4[1]:
                    cell.fill = header_fill; cell.font = white_font
                for _, row in df.iterrows():
                    ws4.append(row.tolist())
                for i, col in enumerate(cols, 1):
                    ws4.column_dimensions[get_column_letter(i)].width = 20

                buf = io.BytesIO()
                wb.save(buf)
                buf.seek(0)
                return buf.getvalue()

            action_rows_xl = []
            for f in sorted(findings, key=lambda x: {'critical':0,'warning':1,'success':2}.get(x['severity'],3)):
                if f.get('action'):
                    priority = "P1 — Immediate" if f['severity']=='critical' else "P2 — This Sprint"
                    action_rows_xl.append({'Priority':priority,'Objective':f['objective'],
                                           'Issue':f['title'],'Action':f['action']})
            for step in chain:
                if step['status'] == 'action':
                    action_rows_xl.append({'Priority':'P3 — Next Sprint','Objective':'Cross-Objective',
                                           'Issue':step['label'],'Action':step['detail']})

            excel_bytes = build_excel(df, findings, chain, score, action_rows_xl)
            st.download_button(
                label="📥 Download Full Excel Report",
                data=excel_bytes,
                file_name="agile_health_report.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        except Exception as e:
            st.error(f"Excel export error: {e}")

    # ══════════════════════════════════════════════════════════════
    # Objective 1: Sprint Completion Forecast
    # ══════════════════════════════════════════════════════════════
    with tabs[1]:
        st.header("📌 Objective 1 — Sprint Completion Forecasting")
        try:
            X1 = df[['Planned_Story_Points_Sprint','Completed_Story_Points','Percent_Done',
                      'Days_Remaining_Sprint','Historical_Velocity','Blocked_Stories','Scope_Change']]
            y1 = df['Success_Label']
            if len(y1.unique()) > 1:
                X1_train, X1_test, y1_train, y1_test = train_test_split(X1, y1, test_size=0.2, random_state=42)
                sprint_model = LogisticRegression(max_iter=1000, class_weight='balanced')
                sprint_model.fit(X1_train, y1_train)
                y1_pred = sprint_model.predict(X1_test)
                st.write(f"✅ Accuracy: {accuracy_score(y1_test, y1_pred):.2f}")
                st.text(classification_report(y1_test, y1_pred))

                st.subheader("🔍 Predict Sprint Success")
                psp = st.number_input("Planned Story Points", 1, 100, 40, key="obj1_psp")
                csp = st.number_input("Completed Story Points", 0, 100, 30, key="obj1_csp")
                percent_done = st.slider("% Done", 0.0, 100.0, 75.0, key="obj1_pd")
                drs = st.number_input("Days Remaining", 0, 30, 5, key="obj1_drs")
                hv  = st.number_input("Historical Velocity", 0, 100, 35, key="obj1_hv")
                bs  = st.number_input("Blocked Stories", 0, 10, 1, key="obj1_bs")
                sc  = st.number_input("Scope Change", -20, 20, 0, key="obj1_sc")

                if st.button("Predict Sprint Success", key="obj1_btn"):
                    features = np.array([[psp, csp, percent_done, drs, hv, bs, sc]])
                    p    = sprint_model.predict(features)[0]
                    prob = sprint_model.predict_proba(features)[0][1]
                    if p:
                        st.success(f"✅ Likely to Complete! ({prob:.2f})")
                    else:
                        st.warning(f"⚠️ Risk of Spillover! ({prob:.2f})")
            else:
                st.error("⚠️ Not enough class variety in Success_Label column.")
        except Exception as e:
            st.error(f"Error in Objective 1: {e}")

    # ══════════════════════════════════════════════════════════════
    # Objective 2: Workload Projection
    # ══════════════════════════════════════════════════════════════
    with tabs[2]:
        st.header("📌 Objective 2 — Workload Projection Forecast")
        try:
            X2 = df[['Planned_Story_Points_Resource','Current_Assigned_SP','Historical_Avg_SP',
                      'Remaining_Days_Resource','High_Priority_Tasks_Resource','Current_Workload_Percent']]
            y2 = df['Expected_Overload']
            if len(y2.unique()) > 1:
                X2_train, X2_test, y2_train, y2_test = train_test_split(X2, y2, test_size=0.2, random_state=42)
                workload_model = RandomForestClassifier()
                workload_model.fit(X2_train, y2_train)
                y2_pred = workload_model.predict(X2_test)
                acc2 = accuracy_score(y2_test, y2_pred)
                st.write(f"✅ Accuracy: {acc2:.2f}")
                if acc2 < 0.60:
                    st.info("ℹ️ **Low predictive signal detected.** The workload features in this dataset have near-zero correlation with the overload label — this is common in synthetic data. In production, richer features (e.g. task completion rate, meeting hours) would improve accuracy significantly.")
                st.text(classification_report(y2_test, y2_pred))

                st.subheader("🔍 Predict Overload Risk")
                psp2 = st.number_input("Planned SP", 1, 100, 35, key="obj2_psp2")
                casp = st.number_input("Current Assigned SP", 0, 100, 40, key="obj2_casp")
                hasp = st.number_input("Historical Avg SP", 1, 100, 30, key="obj2_hasp")
                rdr  = st.number_input("Remaining Days", 1, 30, 5, key="obj2_rdr")
                hpt  = st.number_input("High Priority Tasks", 0, 10, 2, key="obj2_hpt")
                cwp  = st.number_input("Current Workload %", 0, 200, 125, key="obj2_cwp")

                if st.button("Predict Overload", key="obj2_btn"):
                    features = np.array([[psp2, casp, hasp, rdr, hpt, cwp]])
                    pred = workload_model.predict(features)[0]
                    prob = workload_model.predict_proba(features)[0][1]
                    if pred:
                        st.warning(f"⚠️ Overload Risk! ({prob:.2f})")
                    else:
                        st.success(f"✅ Within Capacity ({prob:.2f})")
            else:
                st.error("⚠️ Not enough class variety in Expected_Overload column.")
        except Exception as e:
            st.error(f"Error in Objective 2: {e}")

    # ══════════════════════════════════════════════════════════════
    # Objective 3: Time to Resolve
    # ══════════════════════════════════════════════════════════════
    with tabs[3]:
        st.header("📌 Objective 3 — Time to Resolve Estimation")
        try:
            X3 = pd.get_dummies(df[['Issue_Type','Priority']], drop_first=False)
            X3 = pd.concat([X3, df[['Original_Estimate_Hours','Story_Points_Issue']]], axis=1)
            y3 = df['Resolution_Time_Hours']
            X3_train, X3_test, y3_train, y3_test = train_test_split(X3, y3, test_size=0.2, random_state=42)
            ttr_model = LinearRegression()
            ttr_model.fit(X3_train, y3_train)
            y3_pred = ttr_model.predict(X3_test)
            st.write(f"✅ MSE: {mean_squared_error(y3_test, y3_pred):.2f}")

            st.subheader("🔍 Estimate Time to Resolve")
            issue_type = st.selectbox("Issue Type", ['Bug','Story','Task'], key="obj3_it")
            priority   = st.selectbox("Priority", ['Low','Medium','High'], key="obj3_pri")
            oe = st.number_input("Original Estimate", 1, 50, 8, key="obj3_oe")
            sp = st.number_input("Story Points", 1, 20, 5, key="obj3_sp")

            test_row = pd.DataFrame([{
                'Issue_Type_Bug':   1 if issue_type=='Bug'    else 0,
                'Issue_Type_Story': 1 if issue_type=='Story'  else 0,
                'Issue_Type_Task':  1 if issue_type=='Task'   else 0,
                'Priority_Low':     1 if priority=='Low'      else 0,
                'Priority_Medium':  1 if priority=='Medium'   else 0,
                'Priority_High':    1 if priority=='High'     else 0,
                'Original_Estimate_Hours': oe,
                'Story_Points_Issue': sp
            }])
            test_row = test_row.reindex(columns=X3.columns, fill_value=0)

            if st.button("Estimate Resolution Time", key="obj3_btn"):
                pred_time = max(0, ttr_model.predict(test_row)[0])
                st.info(f"⏰ Estimated Resolution Time: {pred_time:.1f} hours")
        except Exception as e:
            st.error(f"Error in Objective 3: {e}")

    # ══════════════════════════════════════════════════════════════
    # Objective 4: Burnout Risk Alerts
    # ══════════════════════════════════════════════════════════════
    with tabs[4]:
        st.header("📌 Objective 4 — Burnout Risk Alerts")
        try:
            X4 = df[['Total_SP_This_Sprint','Historical_Avg_SP_Burnout',
                      'High_Priority_Tasks_Burnout','Consecutive_Overloads']]
            y4 = df['Risk_Flag']
            if len(y4.unique()) > 1:
                X4_train, X4_test, y4_train, y4_test = train_test_split(X4, y4, test_size=0.2, random_state=42)
                burnout_model = RandomForestClassifier()
                burnout_model.fit(X4_train, y4_train)
                y4_pred = burnout_model.predict(X4_test)
                acc4 = accuracy_score(y4_test, y4_pred)
                st.write(f"✅ Accuracy: {acc4:.2f}")
                if acc4 < 0.60:
                    st.info("ℹ️ **Low predictive signal detected.** Burnout features show near-zero correlation with the risk label in this dataset. Real burnout prediction benefits from additional signals like overtime hours, meeting load, and leave history.")
                st.text(classification_report(y4_test, y4_pred))

                st.subheader("🔍 Check Burnout Risk")
                tsp   = st.number_input("Total SP This Sprint", 0, 100, 40, key="obj4_tsp")
                hasp4 = st.number_input("Historical Avg SP", 1, 100, 25, key="obj4_hasp4")
                hpt4  = st.number_input("High Priority Tasks", 0, 10, 2, key="obj4_hpt4")
                co    = st.number_input("Consecutive Overloads", 0, 5, 2, key="obj4_co")

                if st.button("Check Burnout Risk", key="obj4_btn"):
                    pred = burnout_model.predict([[tsp, hasp4, hpt4, co]])[0]
                    st.warning("⚠️ Burnout Risk Detected!") if pred else st.success("✅ Workload looks healthy!")
            else:
                st.error("⚠️ Not enough class variety in Risk_Flag column.")
        except Exception as e:
            st.error(f"Error in Objective 4: {e}")

    # ══════════════════════════════════════════════════════════════
    # Objective 5: Resource Allocation
    # ══════════════════════════════════════════════════════════════
    with tabs[5]:
        st.header("📌 Objective 5 — Resource Allocation Suggestions")
        try:
            le_summary = LabelEncoder(); le_labels = LabelEncoder()
            df['Summary_enc'] = le_summary.fit_transform(df['Summary'].astype(str))
            df['Labels_enc']  = le_labels.fit_transform(df['Labels'].astype(str))
            X5 = df[['Summary_enc','Labels_enc','Original_Estimate_Resource','Story_Points_Resource']]
            y5 = df['Assignee_Resource']
            X5_train, X5_test, y5_train, y5_test = train_test_split(X5, y5, test_size=0.2, random_state=42)
            alloc_model = RandomForestClassifier()
            alloc_model.fit(X5_train, y5_train)
            y5_pred = alloc_model.predict(X5_test)
            st.write(f"✅ Accuracy: {accuracy_score(y5_test, y5_pred):.2f}")

            st.subheader("🔍 Suggest Assignee")
            summary = st.text_input("Summary (short description)", "Fix bug")
            label   = st.text_input("Label (category)", "Bug")
            oe5 = st.number_input("Original Estimate", 1, 50, 8, key="obj5_oe")
            sp5 = st.number_input("Story Points", 1, 20, 5, key="obj5_sp")

            try: summary_enc = le_summary.transform([summary])[0]
            except: summary_enc = 0
            try: label_enc = le_labels.transform([label])[0]
            except: label_enc = 0

            test_row = pd.DataFrame([{
                'Summary_enc': summary_enc, 'Labels_enc': label_enc,
                'Original_Estimate_Resource': oe5, 'Story_Points_Resource': sp5
            }])

            if st.button("Suggest Assignee", key="obj5_btn"):
                assignee = alloc_model.predict(test_row)[0]
                st.success(f"✅ Recommended Assignee: {assignee}")
        except Exception as e:
            st.error(f"Error in Objective 5: {e}")
